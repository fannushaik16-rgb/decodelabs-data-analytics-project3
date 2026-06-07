import openpyxl
import sqlite3

wb = openpyxl.load_workbook("Dataset_for_Data_Analytics_3.xlsx")
ws = wb.active

conn = sqlite3.connect("orders.db")
cursor = conn.cursor()

cursor.execute("DROP TABLE IF EXISTS orders")

headers = [cell.value for cell in ws[1]]
cols = ", ".join([f'"{h}" TEXT' for h in headers])
cursor.execute(f"CREATE TABLE orders ({cols})")

for row in ws.iter_rows(min_row=2, values_only=True):
    vals = [str(v) if v is not None else None for v in row]
    cursor.execute(f"INSERT INTO orders VALUES ({','.join(['?']*len(headers))})", vals)

conn.commit()
conn.close()
print("✅ Database created: orders.db")